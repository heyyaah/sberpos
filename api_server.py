"""
API Server для управления терминалами SberPOS
Эмулирует серверную часть для тестирования
"""

from flask import Flask, request, jsonify, make_response, render_template_string, send_file
import json
import uuid
import random
import os
import threading
import time
from datetime import datetime, timedelta
from io import BytesIO

# Пробуем импортировать psycopg3, если не получается - работаем без БД
try:
    import psycopg
    from psycopg.types.json import Jsonb
    PSYCOPG_AVAILABLE = True
    print("✅ psycopg3 загружен успешно")
except ImportError as e:
    print(f"⚠️  psycopg не доступен: {e}")
    print("⚠️  Будет использоваться только файловое хранилище")
    PSYCOPG_AVAILABLE = False

# Пробуем импортировать qrcode для генерации QR-кодов
try:
    import qrcode
    QRCODE_AVAILABLE = True
    print("✅ qrcode загружен успешно")
except ImportError:
    print("⚠️  qrcode не доступен, QR-коды не будут генерироваться")
    print("⚠️  Установите: pip install qrcode[pil]")
    QRCODE_AVAILABLE = False

app = Flask(__name__)

# Хранилище данных в памяти
sessions = {}  # session_id -> {terminal_id, authenticated, csrf_token}
terminals = {}  # terminal_id -> {password, current_payload, face_confirm_enabled, owner_id}
device_states = {}  # terminal_id -> {state, amount, last_update}
auto_reset_timers = {}  # terminal_id -> timer
last_seen = {}  # terminal_id -> datetime последнего запроса
users = {}  # user_id -> {username, password, terminals: [], balance: 0}
transactions = {}  # terminal_id -> [{timestamp, amount, type, status}]
shifts = {}  # terminal_id -> {opened_at, closed_at, transactions_count, total_amount}
balance_history = {}  # user_id -> [{timestamp, amount, type, terminal_id, description}]

TERMINALS_FILE = os.environ.get('TERMINALS_FILE', '/data/terminals_db.json') if os.path.exists('/data') else 'terminals_db.json'
USERS_FILE = 'users_db.json'
TRANSACTIONS_FILE = 'transactions_db.json'
SHIFTS_FILE = 'shifts_db.json'
BALANCE_HISTORY_FILE = 'balance_history_db.json'
TERMINAL_TIMEOUT = 10  # секунд без активности для отмены оплаты
DATABASE_URL = os.environ.get('DATABASE_URL')  # PostgreSQL URL от Render

def get_db_connection():
    """Получить подключение к БД"""
    if DATABASE_URL and PSYCOPG_AVAILABLE:
        # Render использует postgres://, но psycopg требует postgresql://
        db_url = DATABASE_URL.replace('postgres://', 'postgresql://')
        return psycopg.connect(db_url)
    return None

def init_db():
    """Инициализация таблиц БД"""
    if not DATABASE_URL or not PSYCOPG_AVAILABLE:
        print("⚠️  DATABASE_URL не найден или psycopg недоступен, используется файловое хранилище")
        return
    
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        
        # Таблица терминалов
        cur.execute('''
            CREATE TABLE IF NOT EXISTS terminals (
                terminal_id VARCHAR(10) PRIMARY KEY,
                data JSONB NOT NULL,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        
        # Таблица пользователей
        cur.execute('''
            CREATE TABLE IF NOT EXISTS users (
                username VARCHAR(100) PRIMARY KEY,
                data JSONB NOT NULL,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        
        # Таблица транзакций
        cur.execute('''
            CREATE TABLE IF NOT EXISTS transactions (
                id SERIAL PRIMARY KEY,
                terminal_id VARCHAR(10) NOT NULL,
                data JSONB NOT NULL,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        
        # Таблица смен
        cur.execute('''
            CREATE TABLE IF NOT EXISTS shifts (
                terminal_id VARCHAR(10) PRIMARY KEY,
                data JSONB NOT NULL,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        
        # Таблица истории баланса
        cur.execute('''
            CREATE TABLE IF NOT EXISTS balance_history (
                id SERIAL PRIMARY KEY,
                user_id VARCHAR(100) NOT NULL,
                data JSONB NOT NULL,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        
        # Таблица избранных терминалов
        cur.execute('''
            CREATE TABLE IF NOT EXISTS favorite_terminals (
                username VARCHAR(100) NOT NULL,
                terminal_id VARCHAR(10) NOT NULL,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                PRIMARY KEY (username, terminal_id)
            )
        ''')
        
        # Таблица сообщений чата поддержки
        cur.execute('''
            CREATE TABLE IF NOT EXISTS support_messages (
                id SERIAL PRIMARY KEY,
                username VARCHAR(100) NOT NULL,
                message TEXT NOT NULL,
                is_admin BOOLEAN DEFAULT FALSE,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        
        # Таблица базы знаний
        cur.execute('''
            CREATE TABLE IF NOT EXISTS knowledge_base (
                id SERIAL PRIMARY KEY,
                title VARCHAR(200) NOT NULL,
                content TEXT NOT NULL,
                category VARCHAR(50) NOT NULL,
                views INT DEFAULT 0,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        
        # Таблица новостей
        cur.execute('''
            CREATE TABLE IF NOT EXISTS news (
                id SERIAL PRIMARY KEY,
                title VARCHAR(200) NOT NULL,
                content TEXT NOT NULL,
                type VARCHAR(50) NOT NULL,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        
        # Таблица фотографий лиц
        cur.execute('''
            CREATE TABLE IF NOT EXISTS face_photos (
                id SERIAL PRIMARY KEY,
                terminal_id VARCHAR(10) NOT NULL,
                uuid VARCHAR(100),
                photo_path VARCHAR(500) NOT NULL,
                timestamp TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                confirmed BOOLEAN DEFAULT FALSE
            )
        ''')
        
        conn.commit()
        cur.close()
        conn.close()
        print("✅ База данных инициализирована (10 таблиц)")
    except Exception as e:
        print(f"❌ Ошибка инициализации БД: {e}")

def auto_reset_to_idle(terminal_id, delay=5):
    """Автоматически сбросить терминал в idle через delay секунд"""
    # Отменяем предыдущий таймер если есть
    if terminal_id in auto_reset_timers:
        auto_reset_timers[terminal_id].cancel()
    
    def reset():
        if terminal_id in terminals:
            terminals[terminal_id]['current_payload'] = {'state': 'idle', 'data': {}}
            terminals[terminal_id]['card_status'] = {
                'pending': True,
                'approved': False
            }
            terminals[terminal_id]['qr_status'] = {
                'pending': True,
                'approved': False
            }
            terminals[terminal_id]['payment_processed'] = False  # Сбрасываем флаг
            device_states[terminal_id] = {
                'state': 'idle',
                'amount': '0',
                'last_update': datetime.now().isoformat()
            }
            print(f"🔄 [AUTO-RESET] {terminal_id} -> idle (all statuses reset)")
    
    timer = threading.Timer(delay, reset)
    timer.start()
    auto_reset_timers[terminal_id] = timer

def load_terminals():
    """Загрузка терминалов из БД или файла"""
    global terminals
    
    # Пробуем загрузить из PostgreSQL
    if DATABASE_URL and PSYCOPG_AVAILABLE:
        try:
            conn = get_db_connection()
            cur = conn.cursor()
            cur.execute('SELECT terminal_id, data FROM terminals')
            rows = cur.fetchall()
            terminals = {row[0]: row[1] for row in rows}
            cur.close()
            conn.close()
            print(f"📂 Загружено {len(terminals)} терминалов из PostgreSQL")
            return
        except Exception as e:
            print(f"❌ Ошибка загрузки из БД: {e}")
    
    # Fallback на файл
    try:
        with open(TERMINALS_FILE, 'r', encoding='utf-8') as f:
            terminals = json.load(f)
            print(f"📂 Загружено {len(terminals)} терминалов из {TERMINALS_FILE}")
    except FileNotFoundError:
        print(f"📂 Файл {TERMINALS_FILE} не найден, создан новый")
        terminals = {}
    except Exception as e:
        print(f"❌ Ошибка загрузки терминалов: {e}")
        terminals = {}

def save_terminals():
    """Сохранение терминалов в БД или файл"""
    # Сохраняем в PostgreSQL
    if DATABASE_URL and PSYCOPG_AVAILABLE:
        try:
            conn = get_db_connection()
            cur = conn.cursor()
            for terminal_id, data in terminals.items():
                cur.execute('''
                    INSERT INTO terminals (terminal_id, data, updated_at)
                    VALUES (%s, %s, CURRENT_TIMESTAMP)
                    ON CONFLICT (terminal_id) 
                    DO UPDATE SET data = %s, updated_at = CURRENT_TIMESTAMP
                ''', (terminal_id, Jsonb(data), Jsonb(data)))
            conn.commit()
            cur.close()
            conn.close()
            return
        except Exception as e:
            print(f"❌ Ошибка сохранения в БД: {e}")
    
    # Fallback на файл
    try:
        with open(TERMINALS_FILE, 'w', encoding='utf-8') as f:
            json.dump(terminals, f, ensure_ascii=False, indent=2)
    except Exception as e:
        print(f"❌ Ошибка сохранения терминалов: {e}")

def load_users():
    """Загрузка пользователей из БД или файла"""
    global users
    
    # Пробуем загрузить из PostgreSQL
    if DATABASE_URL and PSYCOPG_AVAILABLE:
        try:
            conn = get_db_connection()
            cur = conn.cursor()
            cur.execute('SELECT username, data FROM users')
            rows = cur.fetchall()
            users = {row[0]: row[1] for row in rows}
            cur.close()
            conn.close()
            print(f"👥 Загружено {len(users)} пользователей из PostgreSQL")
            return
        except Exception as e:
            print(f"❌ Ошибка загрузки пользователей из БД: {e}")
    
    # Fallback на файл
    try:
        with open(USERS_FILE, 'r', encoding='utf-8') as f:
            users = json.load(f)
            print(f"👥 Загружено {len(users)} пользователей из файла")
    except FileNotFoundError:
        users = {}
        print("👥 Файл пользователей не найден, создан новый")
    except Exception as e:
        print(f"❌ Ошибка загрузки пользователей: {e}")
        users = {}

def save_users():
    """Сохранение пользователей в БД или файл"""
    # Сохраняем в PostgreSQL
    if DATABASE_URL and PSYCOPG_AVAILABLE:
        try:
            conn = get_db_connection()
            cur = conn.cursor()
            for username, data in users.items():
                cur.execute('''
                    INSERT INTO users (username, data, updated_at)
                    VALUES (%s, %s, CURRENT_TIMESTAMP)
                    ON CONFLICT (username) 
                    DO UPDATE SET data = %s, updated_at = CURRENT_TIMESTAMP
                ''', (username, Jsonb(data), Jsonb(data)))
            conn.commit()
            cur.close()
            conn.close()
            return
        except Exception as e:
            print(f"❌ Ошибка сохранения пользователей в БД: {e}")
    
    # Fallback на файл
    try:
        with open(USERS_FILE, 'w', encoding='utf-8') as f:
            json.dump(users, f, ensure_ascii=False, indent=2)
    except Exception as e:
        print(f"❌ Ошибка сохранения пользователей: {e}")

def load_transactions():
    """Загрузка транзакций"""
    global transactions
    try:
        with open(TRANSACTIONS_FILE, 'r', encoding='utf-8') as f:
            transactions = json.load(f)
            print(f"💳 Загружено транзакций для {len(transactions)} терминалов")
    except FileNotFoundError:
        transactions = {}
        print("💳 Файл транзакций не найден, создан новый")
    except Exception as e:
        print(f"❌ Ошибка загрузки транзакций: {e}")
        transactions = {}

def save_transactions():
    """Сохранение транзакций"""
    try:
        with open(TRANSACTIONS_FILE, 'w', encoding='utf-8') as f:
            json.dump(transactions, f, ensure_ascii=False, indent=2)
    except Exception as e:
        print(f"❌ Ошибка сохранения транзакций: {e}")

def load_shifts():
    """Загрузка смен"""
    global shifts
    try:
        with open(SHIFTS_FILE, 'r', encoding='utf-8') as f:
            shifts = json.load(f)
            print(f"📅 Загружено смен для {len(shifts)} терминалов")
    except FileNotFoundError:
        shifts = {}
        print("📅 Файл смен не найден, создан новый")
    except Exception as e:
        print(f"❌ Ошибка загрузки смен: {e}")
        shifts = {}

def save_shifts():
    """Сохранение смен"""
    try:
        with open(SHIFTS_FILE, 'w', encoding='utf-8') as f:
            json.dump(shifts, f, ensure_ascii=False, indent=2)
    except Exception as e:
        print(f"❌ Ошибка сохранения смен: {e}")

def load_balance_history():
    """Загрузка истории баланса"""
    global balance_history
    try:
        with open(BALANCE_HISTORY_FILE, 'r', encoding='utf-8') as f:
            balance_history = json.load(f)
            print(f"💰 Загружено истории баланса для {len(balance_history)} пользователей")
    except FileNotFoundError:
        balance_history = {}
        print("💰 Файл истории баланса не найден, создан новый")
    except Exception as e:
        print(f"❌ Ошибка загрузки истории баланса: {e}")
        balance_history = {}

def save_balance_history():
    """Сохранение истории баланса"""
    try:
        with open(BALANCE_HISTORY_FILE, 'w', encoding='utf-8') as f:
            json.dump(balance_history, f, ensure_ascii=False, indent=2)
    except Exception as e:
        print(f"❌ Ошибка сохранения истории баланса: {e}")

def add_balance(user_id, amount, transaction_type, terminal_id, description):
    """Добавить средства на баланс пользователя"""
    # Находим пользователя по user_id
    username = None
    for uname, udata in users.items():
        if udata.get('user_id') == user_id:
            username = uname
            break
    
    if not username:
        print(f"⚠️  [BALANCE] User {user_id} not found")
        return
    
    # Добавляем к балансу
    users[username]['balance'] = users[username].get('balance', 0) + float(amount)
    
    # Записываем в историю
    if user_id not in balance_history:
        balance_history[user_id] = []
    
    balance_history[user_id].append({
        'timestamp': datetime.now().isoformat(),
        'amount': float(amount),
        'type': transaction_type,
        'terminal_id': terminal_id,
        'description': description
    })
    
    save_users()
    save_balance_history()
    
    print(f"💰 [BALANCE] User {username}: +{amount}₽ (total: {users[username]['balance']}₽)")

def add_transaction(terminal_id, amount, payment_type, status):
    """Добавить транзакцию"""
    if terminal_id not in transactions:
        transactions[terminal_id] = []
    
    transaction = {
        'timestamp': datetime.now().isoformat(),
        'amount': amount,
        'type': payment_type,  # 'card', 'face', 'qr'
        'status': status  # 'success', 'failed'
    }
    
    transactions[terminal_id].append(transaction)
    save_transactions()
    
    # Обновляем статистику смены если смена открыта
    if terminal_id in shifts and shifts[terminal_id].get('opened_at') and not shifts[terminal_id].get('closed_at'):
        shifts[terminal_id]['transactions_count'] = shifts[terminal_id].get('transactions_count', 0) + 1
        if status == 'success':
            shifts[terminal_id]['total_amount'] = shifts[terminal_id].get('total_amount', 0) + float(amount)
        save_shifts()
    
    # Начисляем деньги владельцу терминала если оплата успешна
    if status == 'success' and terminal_id in terminals:
        owner_id = terminals[terminal_id].get('owner_id')
        if owner_id:
            add_balance(owner_id, amount, payment_type, terminal_id, f'Оплата через терминал {terminal_id}')
    
    print(f"💰 [TRANSACTION] {terminal_id}: {amount}₽ via {payment_type} - {status}")

def check_terminal_timeouts():
    """Проверка таймаутов терминалов и отмена оплат"""
    print("🔍 [TIMEOUT CHECKER] Started background thread")
    while True:
        time.sleep(0.5)  # Проверяем каждые 0.5 секунды для быстрой реакции
        now = datetime.now()
        
        for terminal_id in list(last_seen.keys()):
            last_activity = last_seen.get(terminal_id)
            if not last_activity:
                continue
            
            inactive_seconds = (now - last_activity).total_seconds()
            
            # Если терминал не активен больше TERMINAL_TIMEOUT секунд
            if inactive_seconds > TERMINAL_TIMEOUT:
                if terminal_id in terminals:
                    current_state = terminals[terminal_id].get('current_payload', {}).get('state', 'idle')
                    
                    # Если терминал в процессе оплаты - отменяем
                    if current_state in ['pay', 'payPending']:
                        print(f"⏱️  [TIMEOUT] {terminal_id}: no activity for {int(inactive_seconds)}s, cancelling payment (state: {current_state})")
                        terminals[terminal_id]['current_payload'] = {'state': 'idle', 'data': {}}
                        terminals[terminal_id]['card_status'] = {
                            'pending': True,
                            'approved': False
                        }
                        terminals[terminal_id]['payment_processed'] = False
                        if terminal_id in device_states:
                            device_states[terminal_id]['state'] = 'idle'
                            device_states[terminal_id]['last_update'] = now.isoformat()
                    else:
                        print(f"🔍 [TIMEOUT] {terminal_id}: inactive for {int(inactive_seconds)}s but state is {current_state}, skipping")
                
                # Удаляем из отслеживания
                del last_seen[terminal_id]
                print(f"🗑️  [TIMEOUT] {terminal_id}: removed from tracking")

# Инициализация БД
init_db()

# Загрузка терминалов при старте
load_terminals()
load_users()
load_transactions()
load_shifts()
load_balance_history()

# Запуск фонового потока для проверки таймаутов
print("🚀 Starting timeout checker thread...")
timeout_thread = threading.Thread(target=check_terminal_timeouts, daemon=True)
timeout_thread.start()
print("✅ Timeout checker thread started")

def get_session(request):
    """Получить сессию из cookies"""
    session_id = request.cookies.get('session_id')
    return sessions.get(session_id) if session_id else None

def require_auth(f):
    """Декоратор для проверки авторизации"""
    def wrapper(*args, **kwargs):
        session = get_session(request)
        if not session or not session.get('authenticated'):
            return jsonify({'error': 'Unauthorized'}), 401
        return f(session, *args, **kwargs)
    wrapper.__name__ = f.__name__
    return wrapper

@app.route('/api/register_device', methods=['POST'])
def register_terminal():
    """Регистрация нового терминала"""
    data = request.json or {}
    terminal_id = data.get('terminal_id')
    password = data.get('password')
    
    # Валидация формата TRM-####
    if terminal_id:
        if not terminal_id.startswith('TRM-') or len(terminal_id) != 8:
            print(f"❌ [REGISTER] Invalid format: {terminal_id}")
            return jsonify({'error': 'Invalid terminal_id format. Use TRM-####', 'status': 'error'}), 400
        try:
            int(terminal_id[4:])
        except ValueError:
            print(f"❌ [REGISTER] Non-digit ID: {terminal_id}")
            return jsonify({'error': 'Terminal ID must be TRM-#### where #### are digits', 'status': 'error'}), 400
    else:
        # Генерация случайного ID
        terminal_id = f"TRM-{random.randint(1000, 9999)}"
        while terminal_id in terminals:
            terminal_id = f"TRM-{random.randint(1000, 9999)}"
    
    # Валидация пароля
    if password:
        if len(password) != 6 or not password.isdigit():
            print(f"❌ [REGISTER] Invalid password format for {terminal_id}")
            return jsonify({'error': 'Password must be 6 digits', 'status': 'error'}), 400
    else:
        # Генерация случайного пароля
        password = ''.join([str(random.randint(0, 9)) for _ in range(6)])
    
    # Проверка существования
    if terminal_id in terminals:
        print(f"⚠️  [REGISTER] Terminal already exists: {terminal_id}")
        return jsonify({'error': 'Terminal already exists', 'status': 'error'}), 409
    
    # Создание терминала
    terminals[terminal_id] = {
        'password': password,
        'current_payload': {'state': 'idle', 'data': {}},
        'face_confirm_enabled': False,
        'uuid': str(uuid.uuid4()),
        'card_status': {
            'pending': True,
            'approved': False
        },
        'qr_status': {
            'pending': True,
            'approved': False
        },
        'payment_processed': False  # Флаг для предотвращения двойного подтверждения
    }
    
    save_terminals()
    
    print(f"✅ [REGISTER] New terminal created: {terminal_id} / {password}")
    
    return jsonify({
        'success': True,
        'status': 'success',
        'terminal_id': terminal_id,
        'terminal_password': password,
        'uuid': terminals[terminal_id]['uuid']
    }), 201

@app.route('/login', methods=['GET', 'POST'])
def login():
    """Авторизация терминала"""
    if request.method == 'GET':
        return 'Login page', 200
    
    username = request.form.get('username')
    password = request.form.get('password')
    
    if not username or not password:
        print(f"❌ [LOGIN] Missing credentials")
        return jsonify({'error': 'Missing credentials', 'status': 'error'}), 200
    
    terminal = terminals.get(username)
    if not terminal:
        print(f"❌ [LOGIN] Terminal not found: {username}")
        return jsonify({'error': 'Terminal not found', 'status': 'error'}), 200
    
    if terminal['password'] != password:
        print(f"❌ [LOGIN] Wrong password for: {username}")
        return jsonify({'error': 'Invalid password', 'status': 'error'}), 200
    
    # Создаём сессию
    session_id = str(uuid.uuid4())
    csrf_token = str(uuid.uuid4())
    
    sessions[session_id] = {
        'terminal_id': username,
        'authenticated': True,
        'csrf_token': csrf_token
    }
    
    print(f"✅ [LOGIN] Successful login: {username} (session: {session_id[:8]}...)")
    
    response = make_response(jsonify({'success': True, 'status': 'success'}))
    response.set_cookie('session_id', session_id)
    response.set_cookie('csrf', csrf_token)
    
    return response

@app.route('/api/card/status', methods=['GET'])
def card_status():
    """Получить статус карты/оплаты"""
    terminal_id = request.args.get('terminal_id')
    uuid_param = request.args.get('uuid')
    source = request.args.get('source', 'unknown')
    
    if not terminal_id:
        return jsonify({'error': 'Missing terminal_id'}), 400
    
    if terminal_id not in terminals:
        return jsonify({'error': 'Terminal not found'}), 404
    
    terminal = terminals[terminal_id]
    
    # Проверка UUID
    if uuid_param and terminal.get('uuid') != uuid_param:
        return jsonify({'error': 'Invalid UUID'}), 403
    
    # Обновляем время последней активности
    last_seen[terminal_id] = datetime.now()
    
    # Получаем статус подтверждения карты
    card_status_data = terminal.get('card_status', {})
    pending = card_status_data.get('pending', True)
    approved = card_status_data.get('approved', False)
    
    current = terminal.get('current_payload', {'state': 'idle', 'data': {}})
    state = current.get('state', 'idle')
    
    # Если карта приложена (source=sensor) и включен bypass, запускаем автоподтверждение
    if source == 'sensor' and state == 'pay' and pending and not terminal.get('payment_processed', False):
        bypass_card_check = terminal.get('bypass_card_check', False)
        if bypass_card_check and not terminal.get('bypass_timer_started', False):
            terminal['bypass_timer_started'] = True
            current_amount = current.get('data', {}).get('amount', '0')
            
            def auto_confirm():
                print(f"⏱️  [BYPASS] {terminal_id}: Starting 3s countdown (from card sensor)...")
                time.sleep(3)
                if terminal_id in terminals:
                    term = terminals[terminal_id]
                    curr_state = term.get('current_payload', {}).get('state', 'idle')
                    print(f"⏱️  [BYPASS] {terminal_id}: After 3s - state={curr_state}, processed={term.get('payment_processed', False)}")
                    
                    # Проверяем что терминал все еще в pay и оплата не обработана
                    if curr_state == 'pay' and not term.get('payment_processed', False):
                        # Обновляем card_status
                        term['card_status'] = {
                            'pending': False,
                            'approved': True
                        }
                        term['payment_processed'] = True
                        term['bypass_timer_started'] = False
                        add_transaction(terminal_id, current_amount, 'card', 'success')
                        print(f"💳 [BYPASS] {terminal_id}: Auto-approved payment after 3s (bypass enabled)")
                        # Автоматически сбросить в idle через 5 секунд
                        auto_reset_to_idle(terminal_id, delay=5)
                    else:
                        term['bypass_timer_started'] = False
                        print(f"⚠️  [BYPASS] {terminal_id}: Conditions not met - state={curr_state}, processed={term.get('payment_processed', False)}")
            
            threading.Thread(target=auto_confirm, daemon=True).start()
            print(f"⏱️  [BYPASS] {terminal_id}: Auto-confirm timer started (3s) from card sensor")
    
    print(f"📥 [CARD STATUS] {terminal_id}: state={state}, pending={pending}, approved={approved}, source={source}")
    
    return jsonify({
        'success': True,
        'status': state,
        'state': state,
        'pending': pending,
        'approved': approved,
        'data': current.get('data', {}),
        'terminal_id': terminal_id
    }), 200

@app.route('/api/face/status', methods=['GET'])
def face_status():
    """Получить статус оплаты улыбкой"""
    terminal_id = request.args.get('terminal_id')
    uuid_param = request.args.get('uuid')
    
    if not terminal_id:
        return jsonify({'error': 'Missing terminal_id'}), 400
    
    if terminal_id not in terminals:
        return jsonify({'error': 'Terminal not found'}), 404
    
    terminal = terminals[terminal_id]
    
    # Проверка UUID
    if uuid_param and terminal.get('uuid') != uuid_param:
        return jsonify({'error': 'Invalid UUID'}), 403
    
    # Обновляем время последней активности
    last_seen[terminal_id] = datetime.now()
    
    # Получаем статус подтверждения (используем тот же card_status для face)
    face_status_data = terminal.get('card_status', {})
    pending = face_status_data.get('pending', True)
    approved = face_status_data.get('approved', False)
    
    current = terminal.get('current_payload', {'state': 'idle', 'data': {}})
    state = current.get('state', 'idle')
    
    print(f"😊 [FACE STATUS] {terminal_id}: state={state}, pending={pending}, approved={approved}")
    
    return jsonify({
        'success': True,
        'status': state,
        'state': state,
        'pending': pending,
        'approved': approved,
        'data': current.get('data', {}),
        'terminal_id': terminal_id
    }), 200

@app.route('/api/face/upload', methods=['POST'])
def face_upload():
    """Загрузка фото лица для оплаты"""
    try:
        # Получаем данные из формы (multipart/form-data)
        terminal_id = request.form.get('terminal_id')
        uuid_param = request.form.get('uuid')
        
        # Если нет в форме, пробуем из args
        if not terminal_id:
            terminal_id = request.args.get('terminal_id')
        if not uuid_param:
            uuid_param = request.args.get('uuid')
        
        print(f"📸 [FACE UPLOAD] Request from {terminal_id}, content-type: {request.content_type}")
        
        if not terminal_id:
            print(f"❌ [FACE UPLOAD] Missing terminal_id")
            return jsonify({'error': 'Missing terminal_id', 'success': False}), 400
        
        if terminal_id not in terminals:
            print(f"❌ [FACE UPLOAD] Terminal not found: {terminal_id}")
            return jsonify({'error': 'Terminal not found', 'success': False}), 404
        
        terminal = terminals[terminal_id]
        
        # Проверка UUID
        if uuid_param and terminal.get('uuid') != uuid_param:
            print(f"❌ [FACE UPLOAD] Invalid UUID for {terminal_id}")
            return jsonify({'error': 'Invalid UUID', 'success': False}), 403
        
        # Получаем фото если есть
        face_image = None
        image_data = None
        if request.files:
            # Пробуем разные возможные имена поля
            face_image = request.files.get('photo') or request.files.get('face_image') or request.files.get('image') or request.files.get('file')
            if face_image:
                image_data = face_image.read()
                print(f"📸 [FACE UPLOAD] {terminal_id}: received image ({len(image_data)} bytes)")
                
                # Сохраняем фото на диск
                try:
                    # Создаем папку для фото если её нет
                    photos_dir = 'face_photos'
                    os.makedirs(photos_dir, exist_ok=True)
                    
                    # Генерируем имя файла: terminal_id_uuid_timestamp.jpg
                    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                    filename = f"{terminal_id}_{uuid_param}_{timestamp}.jpg"
                    filepath = os.path.join(photos_dir, filename)
                    
                    # Сохраняем файл
                    with open(filepath, 'wb') as f:
                        f.write(image_data)
                    
                    print(f"💾 [FACE UPLOAD] Saved photo to {filepath}")
                    
                    # Сохраняем в базу данных
                    if conn:
                        try:
                            cur = conn.cursor()
                            cur.execute("""
                                INSERT INTO face_photos (terminal_id, uuid, photo_path, timestamp, confirmed)
                                VALUES (%s, %s, %s, NOW(), FALSE)
                            """, (terminal_id, uuid_param, filepath))
                            conn.commit()
                            cur.close()
                            print(f"💾 [FACE UPLOAD] Saved to database")
                        except Exception as db_error:
                            print(f"⚠️ [FACE UPLOAD] Database error: {db_error}")
                            # Продолжаем работу даже если база недоступна
                    
                except Exception as save_error:
                    print(f"⚠️ [FACE UPLOAD] Error saving photo: {save_error}")
        else:
            print(f"📸 [FACE UPLOAD] {terminal_id}: no files in request")
        
        # Проверяем включено ли подтверждение лицом
        face_confirm_enabled = terminal.get('face_confirm_enabled', True)
        
        print(f"😊 [FACE UPLOAD] {terminal_id}: face_confirm_enabled={face_confirm_enabled}, returning success")
        
        # Возвращаем успешный ответ с флагом подтверждения
        return jsonify({
            'success': True,
            'face_confirm_enabled': face_confirm_enabled,
            'terminal_id': terminal_id,
            'message': 'Face uploaded successfully'
        }), 200
        
    except Exception as e:
        print(f"❌ [FACE UPLOAD] Error: {e}")
        return jsonify({'error': str(e), 'success': False}), 500

@app.route('/api/payload', methods=['GET', 'POST'])
@app.route('/admin/set_payload', methods=['POST'])  # Алиас для бота
def payload_handler():
    """Получить или установить payload для терминала"""
    
    if request.method == 'GET':
        # Получение payload - проверка терминала
        terminal_id = request.args.get('terminal_id')
        uuid_param = request.args.get('uuid')
        
        if not terminal_id:
            return jsonify({'error': 'Missing terminal_id'}), 400
        
        if terminal_id not in terminals:
            return jsonify({'error': 'Terminal not found'}), 404
        
        terminal = terminals[terminal_id]
        
        # Проверка UUID
        if uuid_param and terminal.get('uuid') != uuid_param:
            return jsonify({'error': 'Invalid UUID'}), 403
        
        # Обновляем время последней активности
        last_seen[terminal_id] = datetime.now()
        
        current = terminal.get('current_payload', {'state': 'idle', 'data': {}})
        
        response_data = {
            'success': True,
            'state': current.get('state', 'idle'),
            'data': current.get('data', {}),
            'terminal_id': terminal_id
        }
        
        print(f"📥 [GET PAYLOAD] {terminal_id}: state={current.get('state')} (last_seen updated)")
        print(f"   Response: {response_data}")
        
        return jsonify(response_data), 200
    
    else:
        # POST - установка payload (требует авторизации)
        session = get_session(request)
        if not session or not session.get('authenticated'):
            return jsonify({'error': 'Unauthorized'}), 401
        
        data = request.json
        terminal_id = session['terminal_id']
        
        if terminal_id not in terminals:
            return jsonify({'error': 'Terminal not found'}), 404
        
        state = data.get('state', 'idle')
        amount = data.get('amount', '0')
        content = data.get('content', '')
        buttons = data.get('buttons', '')
        
        # Проверяем открыта ли смена если пытаемся отправить оплату
        if state == 'pay':
            owner_id = terminals[terminal_id].get('owner_id')
            bypass_shift_check = terminals[terminal_id].get('bypass_shift_check', False)
            
            if owner_id and not bypass_shift_check:
                # Проверяем есть ли открытая смена
                shift = shifts.get(terminal_id, {})
                if not (shift.get('opened_at') and not shift.get('closed_at')):
                    print(f"❌ [PAYLOAD] {terminal_id}: Cannot send payment - shift is closed")
                    return jsonify({'error': 'Смена закрыта', 'status': 'error'}), 400
        
        terminals[terminal_id]['current_payload'] = {
            'state': state,
            'data': {
                'amount': amount,
                'content': content,
                'buttons': buttons
            }
        }
        
        # Если начинается оплата, устанавливаем pending=True
        if state == 'pay':
            # Генерируем одноразовый пароль для QR-оплаты
            qr_password = ''.join([str(random.randint(0, 9)) for _ in range(6)])
            
            # Обычная логика - требуется подтверждение
            terminals[terminal_id]['card_status'] = {
                'pending': True,
                'approved': False
            }
            terminals[terminal_id]['payment_processed'] = False  # Сбрасываем флаг при новой оплате
            
            terminals[terminal_id]['qr_password'] = qr_password  # Сохраняем пароль
            print(f"🔐 [PAYLOAD] Generated QR password for {terminal_id}: {qr_password}")
        
        device_states[terminal_id] = {
            'state': state,
            'amount': amount,
            'last_update': datetime.now().isoformat()
        }
        
        print(f"📤 [PAYLOAD] Set for {terminal_id}: state={state}, amount={amount}")
        
        return jsonify({'success': True, 'status': 'success'}), 200

@app.route('/admin/set_device_payload', methods=['POST'])
@require_auth
def set_device_payload(session):
    """Установить payload для конкретного устройства"""
    data = request.json
    terminal_id = data.get('terminal_id')
    payload = data.get('payload', 'idle')
    
    if terminal_id not in terminals:
        return jsonify({'error': 'Terminal not found'}), 404
    
    terminals[terminal_id]['current_payload'] = {
        'state': payload,
        'data': {}
    }
    
    device_states[terminal_id] = {
        'state': payload,
        'amount': '0',
        'last_update': datetime.now().isoformat()
    }
    
    return jsonify({'success': True, 'status': 'success'}), 200

@app.route('/admin/set_device_payload_full', methods=['POST'])
@require_auth
def set_device_payload_full(session):
    """Установить полный payload для конкретного устройства (с amount и state)"""
    data = request.json
    terminal_id = data.get('terminal_id')
    state = data.get('state', 'idle')
    amount = data.get('amount', '0')
    
    if terminal_id not in terminals:
        return jsonify({'error': 'Terminal not found'}), 404
    
    # Проверяем открыта ли смена если пытаемся отправить оплату
    if state == 'pay':
        owner_id = terminals[terminal_id].get('owner_id')
        bypass_shift_check = terminals[terminal_id].get('bypass_shift_check', False)
        
        if owner_id and not bypass_shift_check:
            # Проверяем есть ли открытая смена
            shift = shifts.get(terminal_id, {})
            if not (shift.get('opened_at') and not shift.get('closed_at')):
                print(f"❌ [PAYLOAD_FULL] {terminal_id}: Cannot send payment - shift is closed")
                return jsonify({'error': 'Смена закрыта', 'status': 'error'}), 400
    
    terminals[terminal_id]['current_payload'] = {
        'state': state,
        'data': {
            'amount': amount,
            'content': '',
            'buttons': ''
        }
    }
    
    # Если начинается оплата, устанавливаем pending=True
    if state == 'pay':
        # Генерируем одноразовый пароль для QR-оплаты
        qr_password = ''.join([str(random.randint(0, 9)) for _ in range(6)])
        
        # Обычная логика - требуется подтверждение
        terminals[terminal_id]['card_status'] = {
            'pending': True,
            'approved': False
        }
        terminals[terminal_id]['payment_processed'] = False
        
        terminals[terminal_id]['qr_password'] = qr_password
        print(f"🔐 [PAYLOAD_FULL] Generated QR password for {terminal_id}: {qr_password}")
    
    device_states[terminal_id] = {
        'state': state,
        'amount': amount,
        'last_update': datetime.now().isoformat()
    }
    
    print(f"📤 [PAYLOAD_FULL] Set for {terminal_id}: state={state}, amount={amount}")
    
    return jsonify({'success': True, 'status': 'success'}), 200

@app.route('/admin/reset', methods=['POST'])
@require_auth
def reset_all(session):
    """Сбросить все терминалы в idle"""
    for terminal_id in terminals:
        terminals[terminal_id]['current_payload'] = {'state': 'idle', 'data': {}}
        device_states[terminal_id] = {'state': 'idle', 'amount': '0', 'last_update': datetime.now().isoformat()}
    
    return jsonify({'success': True, 'status': 'success'}), 200

@app.route('/admin/set_face_confirm', methods=['POST'])
@require_auth
def set_face_confirm(session):
    """Включить/выключить подтверждение лицом"""
    data = request.json
    terminal_id = data.get('terminal_id')
    enabled = data.get('enabled', False)
    
    if terminal_id not in terminals:
        return jsonify({'error': 'Terminal not found'}), 404
    
    terminals[terminal_id]['face_confirm_enabled'] = enabled
    
    return jsonify({'success': True, 'status': 'success'}), 200

@app.route('/admin/set_bypass_shift_check', methods=['POST'])
@require_auth
def set_bypass_shift_check(session):
    """Включить/выключить обход проверки смены"""
    data = request.json
    terminal_id = data.get('terminal_id')
    enabled = data.get('enabled', False)
    
    if terminal_id not in terminals:
        return jsonify({'error': 'Terminal not found'}), 404
    
    terminals[terminal_id]['bypass_shift_check'] = enabled
    save_terminals()
    
    print(f"🔓 [BYPASS] {terminal_id}: Shift check bypass {'enabled' if enabled else 'disabled'}")
    
    return jsonify({'success': True, 'status': 'success'}), 200

@app.route('/admin/set_bypass_card_check', methods=['POST'])
@require_auth
def set_bypass_card_check(session):
    """Включить/выключить обход проверки карты/лица"""
    data = request.json
    terminal_id = data.get('terminal_id')
    enabled = data.get('enabled', False)
    
    if terminal_id not in terminals:
        return jsonify({'error': 'Terminal not found'}), 404
    
    terminals[terminal_id]['bypass_card_check'] = enabled
    save_terminals()
    
    print(f"💳 [BYPASS] {terminal_id}: Card/face check bypass {'enabled' if enabled else 'disabled'}")
    
    return jsonify({'success': True, 'status': 'success'}), 200

@app.route('/admin/confirm_card', methods=['POST'])
@require_auth
def confirm_card(session):
    """Подтвердить/отклонить карту"""
    data = request.json
    terminal_id = data.get('terminal_id')
    approved = data.get('approved', True)
    
    if terminal_id not in terminals:
        return jsonify({'error': 'Terminal not found'}), 404
    
    # Проверяем что терминал в состоянии оплаты
    current_state = terminals[terminal_id].get('current_payload', {}).get('state', 'idle')
    if current_state not in ['pay', 'payPending']:
        print(f"⚠️  [CARD] {terminal_id}: not in payment state (current: {current_state})")
        return jsonify({'error': 'Not in payment state', 'success': False}), 400
    
    # Проверяем что оплата еще не обработана
    if terminals[terminal_id].get('payment_processed', False):
        print(f"⚠️  [CARD] {terminal_id}: payment already processed")
        return jsonify({'error': 'Payment already processed', 'success': False}), 400
    
    # Помечаем оплату как обработанную
    terminals[terminal_id]['payment_processed'] = True
    
    # Устанавливаем статус подтверждения карты
    terminals[terminal_id]['card_status'] = {
        'pending': False,
        'approved': approved
    }
    
    # Записываем транзакцию
    amount = terminals[terminal_id].get('current_payload', {}).get('data', {}).get('amount', '0')
    add_transaction(terminal_id, amount, 'card', 'success' if approved else 'failed')
    
    if approved:
        print(f"💳 [CARD] {terminal_id}: ✅ approved (pending=False)")
        print(f"   Card status will auto-reset to idle in 5s")
        # Автоматически сбросить в idle через 5 секунд
        auto_reset_to_idle(terminal_id, delay=5)
    else:
        print(f"💳 [CARD] {terminal_id}: ❌ declined (pending=False)")
        # При отклонении показываем экран неудачи и быстрее возвращаемся в idle
        terminals[terminal_id]['current_payload'] = {
            'state': 'paymentFailed',
            'data': {
                'amount': amount,
                'content': 'Оплата отклонена',
                'buttons': ''
            }
        }
        print(f"   Showing paymentFailed screen, will auto-reset to idle in 3s")
        auto_reset_to_idle(terminal_id, delay=3)
    
    return jsonify({'success': True, 'status': 'success'}), 200

@app.route('/admin/confirm_face', methods=['POST'])
@require_auth
def confirm_face(session):
    """Подтвердить/отклонить лицо"""
    data = request.json
    terminal_id = data.get('terminal_id')
    approved = data.get('approved', True)
    
    if terminal_id not in terminals:
        return jsonify({'error': 'Terminal not found'}), 404
    
    # Проверяем что терминал в состоянии оплаты
    current_state = terminals[terminal_id].get('current_payload', {}).get('state', 'idle')
    if current_state not in ['pay', 'payPending']:
        print(f"⚠️  [FACE] {terminal_id}: not in payment state (current: {current_state})")
        return jsonify({'error': 'Not in payment state', 'success': False}), 400
    
    # Проверяем что оплата еще не обработана
    if terminals[terminal_id].get('payment_processed', False):
        print(f"⚠️  [FACE] {terminal_id}: payment already processed")
        return jsonify({'error': 'Payment already processed', 'success': False}), 400
    
    # Помечаем оплату как обработанную
    terminals[terminal_id]['payment_processed'] = True
    
    # Устанавливаем статус подтверждения карты (face использует ту же логику)
    terminals[terminal_id]['card_status'] = {
        'pending': False,
        'approved': approved
    }
    
    # Записываем транзакцию
    amount = terminals[terminal_id].get('current_payload', {}).get('data', {}).get('amount', '0')
    add_transaction(terminal_id, amount, 'face', 'success' if approved else 'failed')
    
    if approved:
        print(f"🙂 [FACE] {terminal_id}: ✅ approved (pending=False)")
        print(f"   Card status will auto-reset to idle in 5s")
        # Автоматически сбросить в idle через 5 секунд
        auto_reset_to_idle(terminal_id, delay=5)
    else:
        print(f"🙂 [FACE] {terminal_id}: ❌ declined (pending=False)")
        # При отклонении показываем экран неудачи и быстрее возвращаемся в idle
        terminals[terminal_id]['current_payload'] = {
            'state': 'paymentFailed',
            'data': {
                'amount': amount,
                'content': 'Оплата отклонена',
                'buttons': ''
            }
        }
        print(f"   Showing paymentFailed screen, will auto-reset to idle in 3s")
        auto_reset_to_idle(terminal_id, delay=3)
    
    return jsonify({'success': True, 'status': 'success'}), 200

@app.route('/admin/status', methods=['GET'])
@require_auth
def get_status(session):
    """Получить статус всех терминалов"""
    devices = []
    
    for terminal_id, terminal_data in terminals.items():
        devices.append({
            'terminal_id': terminal_id,
            'current_payload': terminal_data['current_payload'],
            'face_confirm_enabled': terminal_data.get('face_confirm_enabled', False),
            'last_update': device_states.get(terminal_id, {}).get('last_update', '')
        })
    
    return jsonify({'devices': devices}), 200

@app.route('/admin/face_photos', methods=['GET'])
@require_auth
def get_face_photos(session):
    """Получить список сохранённых фотографий лиц (только для romancev228)"""
    try:
        # Проверка доступа - только для romancev228
        username = session.get('username')
        if username != 'romancev228':
            return jsonify({'error': 'Access denied', 'photos': []}), 403
        
        terminal_id = request.args.get('terminal_id')
        uuid = request.args.get('uuid')
        limit = int(request.args.get('limit', 50))
        
        if not conn:
            return jsonify({'error': 'Database not available', 'photos': []}), 503
        
        cur = conn.cursor()
        
        # Строим запрос с фильтрами
        query = """
            SELECT id, terminal_id, uuid, photo_path, timestamp, confirmed
            FROM face_photos
            WHERE 1=1
        """
        params = []
        
        if terminal_id:
            query += " AND terminal_id = %s"
            params.append(terminal_id)
        
        if uuid:
            query += " AND uuid = %s"
            params.append(uuid)
        
        query += " ORDER BY timestamp DESC LIMIT %s"
        params.append(limit)
        
        cur.execute(query, tuple(params))
        
        photos = []
        for row in cur.fetchall():
            photos.append({
                'id': row[0],
                'terminal_id': row[1],
                'uuid': row[2],
                'photo_path': row[3],
                'timestamp': row[4].isoformat() if row[4] else None,
                'confirmed': row[5]
            })
        
        cur.close()
        return jsonify({'photos': photos, 'success': True}), 200
        
    except Exception as e:
        print(f"❌ Error getting face photos: {e}")
        return jsonify({'error': str(e), 'photos': []}), 500

@app.route('/admin/face_photo/<int:photo_id>', methods=['GET'])
@require_auth
def get_face_photo(session, photo_id):
    """Скачать конкретное фото по ID (только для romancev228)"""
    try:
        # Проверка доступа - только для romancev228
        username = session.get('username')
        if username != 'romancev228':
            return jsonify({'error': 'Access denied'}), 403
        
        if not conn:
            return jsonify({'error': 'Database not available'}), 503
        
        cur = conn.cursor()
        cur.execute("SELECT photo_path FROM face_photos WHERE id = %s", (photo_id,))
        row = cur.fetchone()
        cur.close()
        
        if not row:
            return jsonify({'error': 'Photo not found'}), 404
        
        photo_path = row[0]
        
        if not os.path.exists(photo_path):
            return jsonify({'error': 'Photo file not found on disk'}), 404
        
        return send_file(photo_path, mimetype='image/jpeg')
        
    except Exception as e:
        print(f"❌ Error getting face photo: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/admin/delete_terminal', methods=['POST'])
@require_auth
def delete_terminal(session):
    """Удалить терминал"""
    data = request.json
    terminal_id = data.get('terminal_id')
    
    if not terminal_id:
        return jsonify({'error': 'Missing terminal_id', 'success': False}), 400
    
    if terminal_id not in terminals:
        return jsonify({'error': 'Terminal not found', 'success': False}), 404
    
    # Удаляем терминал
    del terminals[terminal_id]
    
    # Удаляем из БД если используется
    if DATABASE_URL and PSYCOPG_AVAILABLE:
        try:
            conn = get_db_connection()
            cur = conn.cursor()
            cur.execute('DELETE FROM terminals WHERE terminal_id = %s', (terminal_id,))
            conn.commit()
            cur.close()
            conn.close()
        except Exception as e:
            print(f"❌ Ошибка удаления из БД: {e}")
    
    # Удаляем из device_states
    if terminal_id in device_states:
        del device_states[terminal_id]
    
    # Удаляем из last_seen
    if terminal_id in last_seen:
        del last_seen[terminal_id]
    
    # Отменяем таймер если есть
    if terminal_id in auto_reset_timers:
        auto_reset_timers[terminal_id].cancel()
        del auto_reset_timers[terminal_id]
    
    # Сохраняем изменения в файл (fallback)
    save_terminals()
    
    print(f"🗑️  [DELETE] Terminal {terminal_id} deleted")
    
    return jsonify({'success': True, 'status': 'success', 'message': f'Terminal {terminal_id} deleted'}), 200

@app.route('/admin/clear_all_terminals', methods=['POST'])
@require_auth
def clear_all_terminals(session):
    """Удалить ВСЕ терминалы (осторожно!)"""
    global terminals, device_states, last_seen, auto_reset_timers
    
    count = len(terminals)
    
    # Отменяем все таймеры
    for timer in auto_reset_timers.values():
        timer.cancel()
    
    # Очищаем все словари
    terminals.clear()
    device_states.clear()
    last_seen.clear()
    auto_reset_timers.clear()
    
    # Удаляем из БД если используется
    if DATABASE_URL and PSYCOPG_AVAILABLE:
        try:
            conn = get_db_connection()
            cur = conn.cursor()
            cur.execute('DELETE FROM terminals')
            conn.commit()
            cur.close()
            conn.close()
            print(f"🗑️  [CLEAR ALL] Deleted {count} terminals from database")
        except Exception as e:
            print(f"❌ Ошибка очистки БД: {e}")
    
    # Сохраняем пустой файл
    save_terminals()
    
    print(f"🗑️  [CLEAR ALL] Deleted all {count} terminals")
    
    return jsonify({'success': True, 'status': 'success', 'message': f'Deleted {count} terminals'}), 200

@app.route('/api/qr/password', methods=['GET'])
def qr_password():
    """Получить QR-пароль для терминала"""
    terminal_id = request.args.get('terminal_id')
    uuid_param = request.args.get('uuid')
    
    if not terminal_id:
        return jsonify({'error': 'Missing terminal_id'}), 400
    
    if terminal_id not in terminals:
        return jsonify({'error': 'Terminal not found'}), 404
    
    terminal = terminals[terminal_id]
    
    # Проверка UUID
    if uuid_param and terminal.get('uuid') != uuid_param:
        return jsonify({'error': 'Invalid UUID'}), 403
    
    qr_password = terminal.get('qr_password', '')
    
    print(f"🔐 [QR PASSWORD] {terminal_id}: returning password {qr_password}")
    
    return jsonify({
        'success': True,
        'password': qr_password,
        'terminal_id': terminal_id
    }), 200

@app.route('/api/qr/status', methods=['GET'])
def qr_status():
    """Получить статус QR-оплаты"""
    terminal_id = request.args.get('terminal_id')
    uuid_param = request.args.get('uuid')
    
    if not terminal_id:
        return jsonify({'error': 'Missing terminal_id'}), 400
    
    if terminal_id not in terminals:
        return jsonify({'error': 'Terminal not found'}), 404
    
    terminal = terminals[terminal_id]
    
    # Проверка UUID
    if uuid_param and terminal.get('uuid') != uuid_param:
        return jsonify({'error': 'Invalid UUID'}), 403
    
    # Обновляем время последней активности
    last_seen[terminal_id] = datetime.now()
    
    # Получаем статус QR-оплаты
    qr_status_data = terminal.get('qr_status', {})
    pending = qr_status_data.get('pending', True)
    approved = qr_status_data.get('approved', False)
    
    current = terminal.get('current_payload', {'state': 'idle', 'data': {}})
    state = current.get('state', 'idle')
    
    print(f"📱 [QR STATUS] {terminal_id}: state={state}, pending={pending}, approved={approved}")
    
    return jsonify({
        'success': True,
        'status': state,
        'state': state,
        'pending': pending,
        'approved': approved,
        'data': current.get('data', {}),
        'terminal_id': terminal_id
    }), 200

@app.route('/api/qr/confirm', methods=['POST'])
def confirm_qr_public():
    """Публичное подтверждение QR-оплаты с проверкой ключа"""
    data = request.json
    terminal_id = data.get('terminal_id')
    key = data.get('key')
    approved = data.get('approved', True)
    
    if not terminal_id or not key:
        return jsonify({'error': 'Missing terminal_id or key', 'success': False}), 400
    
    if terminal_id not in terminals:
        return jsonify({'error': 'Terminal not found', 'success': False}), 404
    
    # Проверяем ключ
    expected_key = terminals[terminal_id].get('qr_password', '')
    if key != expected_key:
        print(f"❌ [QR CONFIRM] {terminal_id}: Invalid key {key} (expected {expected_key})")
        return jsonify({'error': 'Invalid payment key', 'success': False}), 403
    
    # Проверяем что терминал в состоянии оплаты
    current_state = terminals[terminal_id].get('current_payload', {}).get('state', 'idle')
    if current_state not in ['pay', 'payPending']:
        print(f"⚠️  [QR CONFIRM] {terminal_id}: not in payment state (current: {current_state})")
        return jsonify({'error': 'Not in payment state', 'success': False}), 400
    
    # Проверяем что оплата еще не обработана
    if terminals[terminal_id].get('payment_processed', False):
        print(f"⚠️  [QR CONFIRM] {terminal_id}: payment already processed")
        return jsonify({'error': 'Payment already processed', 'success': False}), 400
    
    # Помечаем оплату как обработанную
    terminals[terminal_id]['payment_processed'] = True
    
    # Устанавливаем статус подтверждения QR
    terminals[terminal_id]['qr_status'] = {
        'pending': False,
        'approved': approved
    }
    
    # Инвалидируем ключ (генерируем новый чтобы старый больше не работал)
    terminals[terminal_id]['qr_password'] = ''.join([str(random.randint(0, 9)) for _ in range(6)])
    
    # Записываем транзакцию
    amount = terminals[terminal_id].get('current_payload', {}).get('data', {}).get('amount', '0')
    add_transaction(terminal_id, amount, 'qr', 'success' if approved else 'failed')
    
    print(f"📱 [QR CONFIRM] {terminal_id}: {'✅ approved' if approved else '❌ declined'} via public API (key: {key})")
    print(f"   QR status will auto-reset to idle in 5s")
    
    # Автоматически сбросить в idle через 5 секунд
    auto_reset_to_idle(terminal_id, delay=5)
    
    return jsonify({'success': True, 'status': 'success'}), 200

@app.route('/admin/confirm_qr', methods=['POST'])
@require_auth
def confirm_qr(session):
    """Подтвердить/отклонить QR-оплату"""
    data = request.json
    terminal_id = data.get('terminal_id')
    approved = data.get('approved', True)
    
    if terminal_id not in terminals:
        return jsonify({'error': 'Terminal not found'}), 404
    
    # Проверяем что терминал в состоянии оплаты
    current_state = terminals[terminal_id].get('current_payload', {}).get('state', 'idle')
    if current_state not in ['pay', 'payPending']:
        print(f"⚠️  [QR] {terminal_id}: not in payment state (current: {current_state})")
        return jsonify({'error': 'Not in payment state', 'success': False}), 400
    
    # Проверяем что оплата еще не обработана
    if terminals[terminal_id].get('payment_processed', False):
        print(f"⚠️  [QR] {terminal_id}: payment already processed")
        return jsonify({'error': 'Payment already processed', 'success': False}), 400
    
    # Помечаем оплату как обработанную
    terminals[terminal_id]['payment_processed'] = True
    
    # Устанавливаем статус подтверждения QR
    terminals[terminal_id]['qr_status'] = {
        'pending': False,
        'approved': approved
    }
    
    # Записываем транзакцию
    amount = terminals[terminal_id].get('current_payload', {}).get('data', {}).get('amount', '0')
    add_transaction(terminal_id, amount, 'qr', 'success' if approved else 'failed')
    
    print(f"📱 [QR] {terminal_id}: {'✅ approved' if approved else '❌ declined'} (pending=False)")
    print(f"   QR status will auto-reset to idle in 5s")
    
    # Автоматически сбросить в idle через 5 секунд
    auto_reset_to_idle(terminal_id, delay=5)
    
    return jsonify({'success': True, 'status': 'success'}), 200


@app.route('/api/qr/initiate', methods=['POST'])
def qr_initiate():
    """Инициировать QR-оплату - переключить терминал на сцену ожидания"""
    data = request.json
    terminal_id = data.get('terminal_id')
    qr_password = data.get('password', '')  # Получаем пароль от клиента
    
    if not terminal_id or terminal_id not in terminals:
        return jsonify({'error': 'Terminal not found', 'success': False}), 404
    
    terminal = terminals[terminal_id]
    
    # Проверяем что терминал в состоянии оплаты
    current_state = terminal.get('current_payload', {}).get('state', 'idle')
    if current_state not in ['pay', 'payPending']:
        return jsonify({'error': 'Not in payment state', 'success': False}), 400
    
    # Сохраняем пароль если передан
    if qr_password:
        terminal['qr_password'] = qr_password
        print(f"🔐 [QR INITIATE] {terminal_id}: Password set to {qr_password}")
    
    # Переключаем терминал на сцену ожидания (payPending)
    current_amount = terminal.get('current_payload', {}).get('data', {}).get('amount', '0')
    terminal['current_payload'] = {
        'state': 'payPending',
        'data': {
            'amount': current_amount,
            'content': '',
            'buttons': ''
        }
    }
    
    # Инициализируем QR статус
    terminal['qr_status'] = {
        'pending': True,
        'approved': False,
        'initiated_at': datetime.now().isoformat()
    }
    
    device_states[terminal_id] = {
        'state': 'payPending',
        'amount': current_amount,
        'last_update': datetime.now().isoformat()
    }
    
    print(f"📱 [QR INITIATE] {terminal_id}: Switched to payPending, waiting for confirmation")
    
    # Проверяем включен ли обход проверки карты/лица
    bypass_card_check = terminal.get('bypass_card_check', False)
    if bypass_card_check:
        # Запускаем таймер автоподтверждения через 3 секунды
        def auto_confirm():
            print(f"⏱️  [BYPASS] {terminal_id}: Starting 3s countdown...")
            time.sleep(3)
            if terminal_id in terminals:
                term = terminals[terminal_id]
                current = term.get('current_payload', {}).get('state', 'idle')
                print(f"⏱️  [BYPASS] {terminal_id}: After 3s - state={current}, processed={term.get('payment_processed', False)}")
                
                # Проверяем что терминал все еще в payPending и оплата не обработана
                if current == 'payPending' and not term.get('payment_processed', False):
                    # Обновляем card_status
                    term['card_status'] = {
                        'pending': False,
                        'approved': True
                    }
                    # Обновляем qr_status (для совместимости)
                    term['qr_status'] = {
                        'pending': False,
                        'approved': True
                    }
                    term['payment_processed'] = True
                    add_transaction(terminal_id, current_amount, 'card', 'success')
                    print(f"💳 [BYPASS] {terminal_id}: Auto-approved payment after 3s (bypass enabled)")
                    # Автоматически сбросить в idle через 5 секунд
                    auto_reset_to_idle(terminal_id, delay=5)
                else:
                    print(f"⚠️  [BYPASS] {terminal_id}: Conditions not met - state={current}, processed={term.get('payment_processed', False)}")
        
        threading.Thread(target=auto_confirm, daemon=True).start()
        print(f"⏱️  [BYPASS] {terminal_id}: Auto-confirm timer started (3s)")
    
    return jsonify({'success': True}), 200

@app.route('/api/qr/check', methods=['GET'])
def qr_check():
    """Проверить статус QR-оплаты (для веб-страницы)"""
    terminal_id = request.args.get('terminal_id')
    
    if not terminal_id or terminal_id not in terminals:
        return jsonify({'error': 'Terminal not found'}), 404
    
    terminal = terminals[terminal_id]
    qr_status_data = terminal.get('qr_status', {'pending': True, 'approved': False})
    
    # Определяем статус для веб-страницы
    if qr_status_data.get('pending', True):
        status = 'pending'
    elif qr_status_data.get('approved', False):
        status = 'success'
    else:
        status = 'failed'
    
    return jsonify({
        'success': True,
        'status': status,
        'pending': qr_status_data.get('pending', True),
        'approved': qr_status_data.get('approved', False)
    }), 200

@app.route('/')
def index():
    return jsonify({
        'status': 'running',
        'endpoints': [
            'POST /api/register_device - Register new terminal',
            'POST /login - Login terminal',
            'GET /api/payload - Get payload',
            'POST /api/payload - Set payload',
            'GET /api/card/status - Get card confirmation status',
            'GET /api/face/status - Get face confirmation status',
            'POST /api/face/upload - Upload face image for payment',
            'GET /api/qr/status - Get QR payment status',
            'POST /api/qr/initiate - Initiate QR payment',
            'GET /api/qr/check - Check QR payment status (web)',
            'GET /p/<terminal_id> - QR payment page',
            'POST /admin/set_device_payload',
            'POST /admin/reset',
            'POST /admin/set_face_confirm',
            'POST /admin/confirm_card',
            'POST /admin/confirm_face',
            'POST /admin/confirm_qr - Confirm/decline QR payment',
            'POST /admin/delete_terminal - Delete terminal',
            'GET /admin/status',
            '--- Personal Cabinet ---',
            'POST /cabinet/register - Register user',
            'POST /cabinet/login - Login user',
            'POST /cabinet/bind_terminal - Bind terminal to user',
            'GET /cabinet/terminals - Get user terminals',
            'GET /cabinet/stats/<terminal_id> - Get terminal statistics',
            'POST /cabinet/shift/open - Open shift',
            'POST /cabinet/shift/close - Close shift',
            'GET /cabinet/shift/status - Get shift status'
        ],
        'terminals_count': len(terminals)
    }), 200

@app.route('/api/qr/generate', methods=['GET'])
def qr_generate():
    """Сгенерировать QR-код для оплаты"""
    terminal_id = request.args.get('terminal_id')
    uuid_param = request.args.get('uuid')
    
    if not terminal_id:
        return jsonify({'error': 'Missing terminal_id'}), 400
    
    if terminal_id not in terminals:
        return jsonify({'error': 'Terminal not found'}), 404
    
    terminal = terminals[terminal_id]
    
    # Проверка UUID
    if uuid_param and terminal.get('uuid') != uuid_param:
        return jsonify({'error': 'Invalid UUID'}), 403
    
    if not QRCODE_AVAILABLE:
        return jsonify({'error': 'QR code generation not available'}), 503
    
    # Получаем текущую сумму оплаты и ключ
    current = terminal.get('current_payload', {'state': 'idle', 'data': {}})
    amount = current.get('data', {}).get('amount', '0')
    qr_key = terminal.get('qr_password', '000000')  # Используем сгенерированный ключ
    
    # Генерируем URL для оплаты с терминалом и ключом
    base_url = request.host_url.rstrip('/')
    payment_url = f"{base_url}/pay/{terminal_id}/key={qr_key}"
    
    # Генерируем QR-код
    qr = qrcode.QRCode(
        version=1,
        error_correction=qrcode.constants.ERROR_CORRECT_L,
        box_size=10,
        border=4,
    )
    qr.add_data(payment_url)
    qr.make(fit=True)
    
    img = qr.make_image(fill_color="black", back_color="white")
    
    # Сохраняем в BytesIO
    img_io = BytesIO()
    img.save(img_io, 'PNG')
    img_io.seek(0)
    
    print(f"🔲 [QR GENERATE] {terminal_id}: Generated QR for {payment_url} (key: {qr_key})")
    
    return send_file(img_io, mimetype='image/png')

# ===== ЛИЧНЫЙ КАБИНЕТ =====

@app.route('/cabinet/register', methods=['POST'])
def cabinet_register():
    """Регистрация пользователя"""
    data = request.json
    username = data.get('username')
    password = data.get('password')
    
    if not username or not password:
        return jsonify({'error': 'Missing username or password'}), 400
    
    if username in users:
        return jsonify({'error': 'User already exists'}), 409
    
    user_id = str(uuid.uuid4())
    users[username] = {
        'user_id': user_id,
        'password': password,
        'terminals': [],
        'balance': 0,  # Начальный баланс
        'created_at': datetime.now().isoformat()
    }
    
    save_users()
    print(f"👤 [REGISTER] New user: {username}")
    
    return jsonify({'success': True, 'user_id': user_id}), 201

@app.route('/cabinet/login', methods=['POST'])
def cabinet_login():
    """Вход в личный кабинет"""
    data = request.json
    username = data.get('username')
    password = data.get('password')
    
    if not username or not password:
        return jsonify({'error': 'Missing credentials'}), 400
    
    user = users.get(username)
    if not user or user['password'] != password:
        return jsonify({'error': 'Invalid credentials'}), 401
    
    # Создаём сессию
    session_id = str(uuid.uuid4())
    sessions[session_id] = {
        'user_id': user['user_id'],
        'username': username,
        'authenticated': True,
        'type': 'user'  # отличаем от терминальной сессии
    }
    
    print(f"👤 [LOGIN] User logged in: {username}")
    
    response = make_response(jsonify({'success': True}))
    response.set_cookie('session_id', session_id)
    
    return response

@app.route('/cabinet/bind_terminal', methods=['POST'])
def cabinet_bind_terminal():
    """Привязать терминал к пользователю"""
    session = get_session(request)
    if not session or session.get('type') != 'user':
        return jsonify({'error': 'Unauthorized'}), 401
    
    data = request.json
    terminal_id = data.get('terminal_id')
    terminal_password = data.get('password')
    
    if not terminal_id or not terminal_password:
        return jsonify({'error': 'Missing terminal_id or password'}), 400
    
    if terminal_id not in terminals:
        return jsonify({'error': 'Terminal not found'}), 404
    
    if terminals[terminal_id]['password'] != terminal_password:
        return jsonify({'error': 'Invalid terminal password'}), 403
    
    # Проверяем что терминал еще не привязан
    if 'owner_id' in terminals[terminal_id]:
        return jsonify({'error': 'Terminal already bound to another user'}), 409
    
    # Привязываем терминал
    username = session['username']
    terminals[terminal_id]['owner_id'] = users[username]['user_id']
    users[username]['terminals'].append(terminal_id)
    
    save_terminals()
    save_users()
    
    print(f"🔗 [BIND] Terminal {terminal_id} bound to user {username}")
    
    return jsonify({'success': True}), 200

@app.route('/cabinet/unbind_terminal', methods=['POST'])
def cabinet_unbind_terminal():
    """Отвязать терминал от пользователя"""
    session = get_session(request)
    if not session or session.get('type') != 'user':
        return jsonify({'error': 'Unauthorized'}), 401
    
    data = request.json
    terminal_id = data.get('terminal_id')
    
    if not terminal_id:
        return jsonify({'error': 'Missing terminal_id'}), 400
    
    username = session['username']
    user_id = users[username]['user_id']
    
    # Проверяем что терминал принадлежит этому пользователю
    if terminal_id not in users[username]['terminals']:
        return jsonify({'error': 'Terminal not owned by user'}), 403
    
    # Отвязываем терминал
    if terminal_id in terminals:
        terminals[terminal_id].pop('owner_id', None)
        save_terminals()
    
    users[username]['terminals'].remove(terminal_id)
    save_users()
    
    print(f"🔓 [UNBIND] Terminal {terminal_id} unbound from user {username}")
    
    return jsonify({'success': True}), 200

@app.route('/admin/force_unbind_terminal', methods=['POST'])
@require_auth
def admin_force_unbind_terminal(session):
    """Принудительно отвязать терминал (для админов)"""
    data = request.json
    terminal_id = data.get('terminal_id')
    
    if not terminal_id:
        return jsonify({'error': 'Missing terminal_id'}), 400
    
    if terminal_id not in terminals:
        return jsonify({'error': 'Terminal not found'}), 404
    
    # Находим владельца
    owner_id = terminals[terminal_id].get('owner_id')
    if owner_id:
        # Удаляем из списка терминалов пользователя
        for username, user_data in users.items():
            if user_data.get('user_id') == owner_id and terminal_id in user_data.get('terminals', []):
                user_data['terminals'].remove(terminal_id)
                save_users()
                break
    
    # Отвязываем терминал
    terminals[terminal_id].pop('owner_id', None)
    save_terminals()
    
    print(f"🔓 [FORCE UNBIND] Terminal {terminal_id} force unbound")
    
    return jsonify({'success': True, 'message': f'Terminal {terminal_id} unbound'}), 200

@app.route('/cabinet/terminals', methods=['GET'])
def cabinet_terminals():
    """Получить список терминалов пользователя"""
    session = get_session(request)
    if not session or session.get('type') != 'user':
        return jsonify({'error': 'Unauthorized'}), 401
    
    username = session['username']
    user_terminals = users[username]['terminals']
    favorites = users[username].get('favorites', [])
    
    terminals_data = []
    for terminal_id in user_terminals:
        if terminal_id in terminals:
            terminal = terminals[terminal_id]
            shift_status = shifts.get(terminal_id, {})
            
            terminals_data.append({
                'terminal_id': terminal_id,
                'current_state': terminal.get('current_payload', {}).get('state', 'idle'),
                'shift_opened': bool(shift_status.get('opened_at') and not shift_status.get('closed_at')),
                'shift_transactions': shift_status.get('transactions_count', 0),
                'shift_total': shift_status.get('total_amount', 0),
                'is_favorite': terminal_id in favorites
            })
    
    return jsonify({'terminals': terminals_data, 'favorites': favorites}), 200

@app.route('/cabinet/stats/<terminal_id>', methods=['GET'])
def cabinet_stats(terminal_id):
    """Получить статистику терминала"""
    session = get_session(request)
    if not session or session.get('type') != 'user':
        return jsonify({'error': 'Unauthorized'}), 401
    
    username = session['username']
    
    # Проверяем что терминал принадлежит пользователю
    if terminal_id not in users[username]['terminals']:
        return jsonify({'error': 'Access denied'}), 403
    
    if terminal_id not in terminals:
        return jsonify({'error': 'Terminal not found'}), 404
    
    # Получаем транзакции
    terminal_transactions = transactions.get(terminal_id, [])
    
    # Статистика
    total_transactions = len(terminal_transactions)
    successful_transactions = len([t for t in terminal_transactions if t['status'] == 'success'])
    total_amount = sum(float(t['amount']) for t in terminal_transactions if t['status'] == 'success')
    
    # Группировка по типам оплаты
    by_type = {}
    for t in terminal_transactions:
        payment_type = t['type']
        if payment_type not in by_type:
            by_type[payment_type] = {'count': 0, 'amount': 0}
        by_type[payment_type]['count'] += 1
        if t['status'] == 'success':
            by_type[payment_type]['amount'] += float(t['amount'])
    
    # Статус смены
    shift = shifts.get(terminal_id, {})
    
    return jsonify({
        'terminal_id': terminal_id,
        'total_transactions': total_transactions,
        'successful_transactions': successful_transactions,
        'failed_transactions': total_transactions - successful_transactions,
        'total_amount': total_amount,
        'by_type': by_type,
        'shift': {
            'opened': bool(shift.get('opened_at') and not shift.get('closed_at')),
            'opened_at': shift.get('opened_at'),
            'closed_at': shift.get('closed_at'),
            'transactions_count': shift.get('transactions_count', 0),
            'total_amount': shift.get('total_amount', 0)
        },
        'recent_transactions': terminal_transactions[-10:]  # последние 10
    }), 200

@app.route('/cabinet/shift/open', methods=['POST'])
def cabinet_shift_open():
    """Открыть смену"""
    session = get_session(request)
    if not session or session.get('type') != 'user':
        return jsonify({'error': 'Unauthorized'}), 401
    
    data = request.json
    terminal_id = data.get('terminal_id')
    
    username = session['username']
    if terminal_id not in users[username]['terminals']:
        return jsonify({'error': 'Access denied'}), 403
    
    # Проверяем что смена не открыта
    if terminal_id in shifts and shifts[terminal_id].get('opened_at') and not shifts[terminal_id].get('closed_at'):
        return jsonify({'error': 'Shift already opened'}), 400
    
    shifts[terminal_id] = {
        'opened_at': datetime.now().isoformat(),
        'closed_at': None,
        'transactions_count': 0,
        'total_amount': 0
    }
    
    save_shifts()
    print(f"📅 [SHIFT] Opened for {terminal_id}")
    
    return jsonify({'success': True}), 200

@app.route('/cabinet/shift/close', methods=['POST'])
def cabinet_shift_close():
    """Закрыть смену"""
    session = get_session(request)
    if not session or session.get('type') != 'user':
        return jsonify({'error': 'Unauthorized'}), 401
    
    data = request.json
    terminal_id = data.get('terminal_id')
    
    username = session['username']
    if terminal_id not in users[username]['terminals']:
        return jsonify({'error': 'Access denied'}), 403
    
    # Проверяем что смена открыта
    if terminal_id not in shifts or not shifts[terminal_id].get('opened_at') or shifts[terminal_id].get('closed_at'):
        return jsonify({'error': 'No open shift'}), 400
    
    shifts[terminal_id]['closed_at'] = datetime.now().isoformat()
    save_shifts()
    
    print(f"📅 [SHIFT] Closed for {terminal_id}")
    
    return jsonify({
        'success': True,
        'shift': shifts[terminal_id]
    }), 200

@app.route('/cabinet/shift/status', methods=['GET'])
def cabinet_shift_status():
    """Получить статус смены"""
    session = get_session(request)
    if not session or session.get('type') != 'user':
        return jsonify({'error': 'Unauthorized'}), 401
    
    terminal_id = request.args.get('terminal_id')
    
    username = session['username']
    if terminal_id not in users[username]['terminals']:
        return jsonify({'error': 'Access denied'}), 403
    
    shift = shifts.get(terminal_id, {})
    
    return jsonify({
        'opened': bool(shift.get('opened_at') and not shift.get('closed_at')),
        'shift': shift
    }), 200

@app.route('/cabinet/balance', methods=['GET'])
def cabinet_balance():
    """Получить баланс пользователя"""
    session = get_session(request)
    if not session or session.get('type') != 'user':
        return jsonify({'error': 'Unauthorized'}), 401
    
    username = session['username']
    user = users[username]
    user_id = user['user_id']
    
    # История баланса
    history = balance_history.get(user_id, [])
    
    return jsonify({
        'balance': user.get('balance', 0),
        'history': history[-20:]  # последние 20 операций
    }), 200

@app.route('/cabinet/terminal/<terminal_id>/favorite', methods=['POST'])
def toggle_favorite(terminal_id):
    """Добавить/удалить терминал из избранного"""
    session = get_session(request)
    if not session or session.get('type') != 'user':
        return jsonify({'error': 'Unauthorized'}), 401
    
    username = session['username']
    
    # Проверяем что терминал принадлежит пользователю
    if terminal_id not in users[username]['terminals']:
        return jsonify({'error': 'Access denied'}), 403
    
    # Получаем текущий список избранных
    favorites = users[username].get('favorites', [])
    
    # Переключаем статус
    if terminal_id in favorites:
        favorites.remove(terminal_id)
        is_favorite = False
    else:
        favorites.append(terminal_id)
        is_favorite = True
    
    users[username]['favorites'] = favorites
    save_users()
    
    print(f"⭐ [FAVORITE] {username}: {terminal_id} {'added to' if is_favorite else 'removed from'} favorites")
    
    return jsonify({'success': True, 'is_favorite': is_favorite}), 200

@app.route('/cabinet/analytics', methods=['GET'])
def cabinet_analytics():
    """Получить аналитику по всем терминалам пользователя"""
    session = get_session(request)
    if not session or session.get('type') != 'user':
        return jsonify({'error': 'Unauthorized'}), 401
    
    username = session['username']
    user_terminals = users[username]['terminals']
    
    period = request.args.get('period', 'day')  # day, week, month
    date_from = request.args.get('date_from')
    date_to = request.args.get('date_to')
    
    # Собираем все транзакции пользователя
    all_transactions = []
    for terminal_id in user_terminals:
        if terminal_id in transactions:
            for t in transactions[terminal_id]:
                t_copy = t.copy()
                t_copy['terminal_id'] = terminal_id
                all_transactions.append(t_copy)
    
    # Фильтруем по датам если указаны
    if date_from:
        all_transactions = [t for t in all_transactions if t['timestamp'] >= date_from]
    if date_to:
        all_transactions = [t for t in all_transactions if t['timestamp'] <= date_to + 'T23:59:59']
    
    # Группируем по периодам
    from collections import defaultdict
    grouped = defaultdict(lambda: {'success': 0, 'failed': 0, 'amount': 0})
    
    for t in all_transactions:
        dt = datetime.fromisoformat(t['timestamp'])
        
        if period == 'day':
            key = dt.strftime('%Y-%m-%d')
        elif period == 'week':
            key = dt.strftime('%Y-W%W')
        else:  # month
            key = dt.strftime('%Y-%m')
        
        if t['status'] == 'success':
            grouped[key]['success'] += 1
            grouped[key]['amount'] += float(t['amount'])
        else:
            grouped[key]['failed'] += 1
    
    # Топ терминалов по выручке
    terminal_revenue = defaultdict(lambda: {'amount': 0, 'count': 0})
    for t in all_transactions:
        if t['status'] == 'success':
            terminal_revenue[t['terminal_id']]['amount'] += float(t['amount'])
            terminal_revenue[t['terminal_id']]['count'] += 1
    
    top_terminals = sorted(
        [{'terminal_id': k, 'amount': v['amount'], 'count': v['count']} 
         for k, v in terminal_revenue.items()],
        key=lambda x: x['amount'],
        reverse=True
    )[:10]
    
    # Конверсия по типам оплаты
    by_type = defaultdict(lambda: {'success': 0, 'failed': 0})
    for t in all_transactions:
        by_type[t['type']][t['status']] += 1
    
    conversion = {}
    for payment_type, stats in by_type.items():
        total = stats['success'] + stats['failed']
        conversion[payment_type] = {
            'rate': (stats['success'] / total * 100) if total > 0 else 0,
            'success': stats['success'],
            'failed': stats['failed']
        }
    
    return jsonify({
        'chart_data': dict(grouped),
        'top_terminals': top_terminals,
        'conversion': conversion
    }), 200

@app.route('/cabinet/export/transactions', methods=['GET'])
def export_transactions():
    """Экспорт транзакций в Excel или CSV"""
    session = get_session(request)
    if not session or session.get('type') != 'user':
        return jsonify({'error': 'Unauthorized'}), 401
    
    username = session['username']
    user_terminals = users[username]['terminals']
    export_format = request.args.get('format', 'csv')  # csv или xlsx
    terminal_id = request.args.get('terminal_id')  # опционально
    
    # Собираем транзакции
    all_transactions = []
    if terminal_id and terminal_id in user_terminals:
        # Экспорт одного терминала
        if terminal_id in transactions:
            for t in transactions[terminal_id]:
                t_copy = t.copy()
                t_copy['terminal_id'] = terminal_id
                all_transactions.append(t_copy)
    else:
        # Экспорт всех терминалов
        for tid in user_terminals:
            if tid in transactions:
                for t in transactions[tid]:
                    t_copy = t.copy()
                    t_copy['terminal_id'] = tid
                    all_transactions.append(t_copy)
    
    # Сортируем по дате
    all_transactions.sort(key=lambda x: x['timestamp'], reverse=True)
    
    if export_format == 'xlsx':
        # Excel экспорт
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Транзакции"
            
            # Заголовки
            headers = ['Дата и время', 'Терминал', 'Сумма (₽)', 'Тип оплаты', 'Статус']
            ws.append(headers)
            
            # Стиль заголовков
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
            
            # Данные
            for t in all_transactions:
                ws.append([
                    t['timestamp'],
                    t['terminal_id'],
                    float(t['amount']),
                    t['type'],
                    'Успешно' if t['status'] == 'success' else 'Неудачно'
                ])
            
            # Сохраняем в память
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            response = make_response(output.read())
            response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            response.headers['Content-Disposition'] = f'attachment; filename=transactions_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
            
            return response
        except ImportError:
            return jsonify({'error': 'openpyxl not installed'}), 500
    
    else:
        # CSV экспорт
        import csv
        
        output = BytesIO()
        writer = csv.writer(output)
        
        # Заголовки
        writer.writerow(['Дата и время', 'Терминал', 'Сумма (₽)', 'Тип оплаты', 'Статус'])
        
        # Данные
        for t in all_transactions:
            writer.writerow([
                t['timestamp'],
                t['terminal_id'],
                t['amount'],
                t['type'],
                'Успешно' if t['status'] == 'success' else 'Неудачно'
            ])
        
        output.seek(0)
        response = make_response(output.read().decode('utf-8'))
        response.headers['Content-Type'] = 'text/csv; charset=utf-8'
        response.headers['Content-Disposition'] = f'attachment; filename=transactions_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
        
        return response

@app.route('/cabinet/support/messages', methods=['GET'])
def get_support_messages():
    """Получить сообщения чата поддержки"""
    session = get_session(request)
    if not session or session.get('type') != 'user':
        return jsonify({'error': 'Unauthorized'}), 401
    
    username = session['username']
    user = users.get(username)
    
    # Проверяем блокировку
    try:
        with open('users.json', 'r') as f:
            bot_users = json.load(f)
            telegram_id = user.get('telegram_id')
            if telegram_id and telegram_id in bot_users.get('blocked', []):
                return jsonify({'error': 'Вы заблокированы в боте', 'blocked': True}), 403
    except Exception:
        pass
    
    # Загружаем сообщения из заявки
    messages = []
    try:
        with open('tickets.json', 'r') as f:
            tickets = json.load(f)
            telegram_id = user.get('telegram_id')
            
            # Ищем заявку пользователя
            for tid, ticket in tickets.items():
                if ticket.get('user_id') == telegram_id or ticket.get('username') == username:
                    # Добавляем описание как первое сообщение
                    messages.append({
                        'message': ticket.get('description', ''),
                        'is_admin': False,
                        'timestamp': datetime.fromtimestamp(ticket.get('created_at', 0)).isoformat()
                    })
                    
                    # Добавляем все сообщения из истории
                    for msg in ticket.get('messages', []):
                        messages.append({
                            'message': msg.get('text', ''),
                            'is_admin': msg.get('from') == 'admin',
                            'timestamp': datetime.fromtimestamp(msg.get('timestamp', 0)).isoformat()
                        })
                    break
    except Exception as e:
        print(f"❌ Ошибка загрузки заявок: {e}")
    
    # Также загружаем из БД (для совместимости)
    if DATABASE_URL and PSYCOPG_AVAILABLE:
        try:
            conn = get_db_connection()
            cur = conn.cursor()
            cur.execute('''
                SELECT message, is_admin, created_at 
                FROM support_messages 
                WHERE username = %s 
                ORDER BY created_at ASC
            ''', (username,))
            rows = cur.fetchall()
            db_messages = [{'message': r[0], 'is_admin': r[1], 'timestamp': r[2].isoformat()} for r in rows]
            
            # Объединяем сообщения (избегаем дубликатов)
            for db_msg in db_messages:
                if not any(m['message'] == db_msg['message'] and m['timestamp'] == db_msg['timestamp'] for m in messages):
                    messages.append(db_msg)
            
            cur.close()
            conn.close()
        except Exception as e:
            print(f"❌ Ошибка загрузки из БД: {e}")
    
    # Сортируем по времени
    messages.sort(key=lambda x: x['timestamp'])
    
    return jsonify({'messages': messages}), 200

@app.route('/cabinet/support/send', methods=['POST'])
def send_support_message():
    """Отправить сообщение в поддержку"""
    session = get_session(request)
    if not session or session.get('type') != 'user':
        return jsonify({'error': 'Unauthorized'}), 401
    
    username = session['username']
    data = request.json
    message = data.get('message', '').strip()
    
    if not message:
        return jsonify({'error': 'Empty message'}), 400
    
    # Получаем user_id пользователя
    user = users.get(username)
    if not user:
        return jsonify({'error': 'User not found'}), 404
    
    # Проверяем заблокирован ли пользователь в боте
    try:
        import requests
        BOT_TOKEN = os.environ.get('SUPPORT_BOT_TOKEN', '8742060687:AAHfsRrh1EKju-ZZ3CLHb2cBl-CkM63fByc')
        
        # Проверяем блокировку через файл users.json бота
        try:
            with open('users.json', 'r') as f:
                bot_users = json.load(f)
                # Ищем telegram_id пользователя по username
                telegram_id = user.get('telegram_id')
                if telegram_id and telegram_id in bot_users.get('blocked', []):
                    return jsonify({'error': 'Вы заблокированы в боте'}), 403
        except Exception:
            pass
        
        # Загружаем или создаем заявку
        try:
            with open('tickets.json', 'r') as f:
                tickets = json.load(f)
        except Exception:
            tickets = {}
        
        # Ищем открытую заявку пользователя или создаем новую
        user_ticket_id = None
        telegram_id = user.get('telegram_id')
        
        if telegram_id:
            for tid, ticket in tickets.items():
                if ticket.get('user_id') == telegram_id and ticket.get('status') in ['open', 'in_progress']:
                    user_ticket_id = tid
                    break
        
        if not user_ticket_id:
            # Создаем новую заявку
            ticket_id = max([int(k) for k in tickets.keys()], default=0) + 1
            user_ticket_id = str(ticket_id)
            
            tickets[user_ticket_id] = {
                'user_id': telegram_id or 0,
                'username': username,
                'first_name': username,
                'description': message,
                'status': 'open',
                'created_at': time.time(),
                'messages': [],
                'from_web': True
            }
            
            print(f"🆕 [SUPPORT] Created new ticket #{user_ticket_id} from web for {username}")
            
            # Уведомляем админов через Telegram с кнопками
            try:
                with open('users.json', 'r') as f:
                    bot_users = json.load(f)
                    admins = bot_users.get('admins', [])
                    
                    print(f"📢 [SUPPORT] Notifying {len(admins)} admins about ticket #{user_ticket_id}")
                    
                    # Создаем inline клавиатуру с кнопками
                    keyboard = {
                        'inline_keyboard': [
                            [{'text': '✅ Принять', 'callback_data': f'accept_{user_ticket_id}'}],
                        ]
                    }
                    
                    for admin_id in admins:
                        try:
                            response = requests.post(
                                f'https://api.telegram.org/bot{BOT_TOKEN}/sendMessage',
                                json={
                                    'chat_id': admin_id,
                                    'text': f"🆕 🌐 Заявка #{user_ticket_id}\n"
                                           f"От: {username} (веб-кабинет)\n"
                                           f"Статус: open\n"
                                           f"Создана: {datetime.now().strftime('%d.%m.%Y %H:%M')}\n\n"
                                           f"📝 {message}",
                                    'reply_markup': keyboard
                                }
                            )
                            if response.status_code == 200:
                                print(f"✅ [SUPPORT] Notified admin {admin_id}")
                            else:
                                print(f"❌ [SUPPORT] Failed to notify admin {admin_id}: {response.text}")
                        except Exception as e:
                            print(f"❌ [SUPPORT] Error notifying admin {admin_id}: {e}")
            except Exception as e:
                print(f"❌ [SUPPORT] Error loading admins: {e}")
        else:
            # Добавляем сообщение к существующей заявке
            tickets[user_ticket_id]['messages'].append({
                'from': 'user',
                'user_id': telegram_id or 0,
                'text': message,
                'timestamp': time.time(),
                'from_web': True
            })
            
            # Уведомляем назначенного админа
            assigned_to = tickets[user_ticket_id].get('assigned_to')
            if assigned_to:
                try:
                    requests.post(
                        f'https://api.telegram.org/bot{BOT_TOKEN}/sendMessage',
                        json={
                            'chat_id': assigned_to,
                            'text': f"💬 {username} ответил на заявку #{user_ticket_id} (из веб-кабинета):\n\n{message}"
                        }
                    )
                except Exception:
                    pass
        
        # Сохраняем заявки
        with open('tickets.json', 'w', encoding='utf-8') as f:
            json.dump(tickets, f, ensure_ascii=False, indent=2)
        
        print(f"💾 [SUPPORT] Saved ticket #{user_ticket_id} to tickets.json")
        
        # Сохраняем в БД для отображения в веб-интерфейсе
        if DATABASE_URL and PSYCOPG_AVAILABLE:
            try:
                conn = get_db_connection()
                cur = conn.cursor()
                cur.execute('''
                    INSERT INTO support_messages (username, message, is_admin, created_at)
                    VALUES (%s, %s, FALSE, CURRENT_TIMESTAMP)
                ''', (username, message))
                conn.commit()
                cur.close()
                conn.close()
            except Exception as e:
                print(f"❌ Ошибка сохранения в БД: {e}")
        
        print(f"💬 [SUPPORT] {username}: {message} (ticket #{user_ticket_id})")
        
        return jsonify({'success': True, 'ticket_id': user_ticket_id}), 200
        
    except Exception as e:
        print(f"❌ Ошибка отправки в бот: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/cabinet/faq', methods=['GET'])
def get_faq():
    """Получить FAQ"""
    # Простой статический FAQ
    faq = [
        {
            'question': 'Как привязать терминал?',
            'answer': 'Введите ID терминала (TRM-####) и пароль в разделе "Привязать терминал".'
        },
        {
            'question': 'Как открыть смену?',
            'answer': 'Выберите терминал и нажмите "Открыть смену" в разделе управления сменой.'
        },
        {
            'question': 'Как экспортировать транзакции?',
            'answer': 'Перейдите в раздел "Аналитика" и нажмите кнопку "Скачать Excel" или "Скачать CSV".'
        },
        {
            'question': 'Что делать если терминал не отвечает?',
            'answer': 'Проверьте интернет-соединение терминала. Если проблема сохраняется, обратитесь в поддержку через чат.'
        },
        {
            'question': 'Как посмотреть статистику?',
            'answer': 'Нажмите "Подробнее" на карточке терминала или перейдите в раздел "Аналитика" для общей статистики.'
        }
    ]
    return jsonify({'faq': faq}), 200

@app.route('/cabinet/news', methods=['GET'])
def get_news():
    """Получить новости"""
    # Простые статические новости
    news = [
        {
            'id': 1,
            'title': 'Добавлена темная тема',
            'content': 'Теперь вы можете переключаться между светлой и темной темой интерфейса.',
            'type': 'feature',
            'date': '2026-04-06'
        },
        {
            'id': 2,
            'title': 'Новая аналитика',
            'content': 'Доступны графики транзакций, рейтинг терминалов и анализ конверсии.',
            'type': 'feature',
            'date': '2026-04-06'
        },
        {
            'id': 3,
            'title': 'Экспорт в Excel и CSV',
            'content': 'Теперь можно экспортировать транзакции в удобном формате для анализа.',
            'type': 'feature',
            'date': '2026-04-06'
        },
        {
            'id': 4,
            'title': 'Чат с поддержкой',
            'content': 'Добавлен виджет чата для быстрой связи с технической поддержкой.',
            'type': 'feature',
            'date': '2026-04-06'
        }
    ]
    return jsonify({'news': news}), 200

@app.route('/static/logo.jpg')
def serve_logo():
    """Отдать логотип"""
    try:
        return send_file('static_logo.jpg', mimetype='image/jpeg')
    except Exception:
        return '', 404


if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5001))
    print(f"🚀 API Server запущен на порту {port}")
    print(f"📌 Загружено терминалов: {len(terminals)}")
    print("   Регистрация: POST /register")
    print("   Формат: TRM-#### (4 цифры), пароль: 6 цифр")
    app.run(host='0.0.0.0', port=port, debug=False)
